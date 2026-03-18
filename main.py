"""
Telegram Bot with Mini App - NytroBC
"""

import asyncio
import io
from aiogram import Bot, Dispatcher, Router, types, F
from aiogram.filters import Command, CommandObject
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    ReplyKeyboardMarkup,
    KeyboardButton,
    WebAppInfo,
    FSInputFile,
    BufferedInputFile,
    LabeledPrice,
)
import os

# Токен вашего бота (замените на свой)
BOT_TOKEN = "8546925390:AAHbsYxLq1OBOLH2SbBsxYYsF58f22nayjM"

# Словарь обязательных каналов для подписки
REQUIRED_CHANNELS = {"NytroBC channel": "@nbc_channeI"}

# URL вашего Mini App
MINI_APP_URL = (
    "https://9617e0e6-f04e-4e15-900a-b69b0a626052-00-2y89sdzkr0mmy.sisko.replit.dev"
)

# Инициализация бота и диспетчера
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())
router = Router()


class ExportState(StatesGroup):
    waiting_start_date = State()
    waiting_end_date = State()


# URL изображений (локальные файлы)
START_IMAGE_PATH = os.path.join(
    os.path.dirname(__file__), "static", "images", "Banner.png"
)
LANGUAGE_IMAGE_PATH = os.path.join(
    os.path.dirname(__file__), "static", "images", "language.jpg"
)


def build_main_menu_keyboard(user_id, mini_app_url, lang="ru"):
    """Строит главное инлайн-меню на нужном языке"""
    if lang == "en":
        return InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=" My Cabinet",
                        style="success",
                        icon_custom_emoji_id="5282843764451195532",
                        web_app=WebAppInfo(url=mini_app_url),
                    )
                ],
                [
                    InlineKeyboardButton(
                        text=" Subscription",
                        callback_data=f"sub_{user_id}",
                        style="primary",
                        icon_custom_emoji_id="5438496463044752972",
                    ),
                    InlineKeyboardButton(
                        text=" Support",
                        callback_data=f"sup_{user_id}",
                        style="primary",
                        icon_custom_emoji_id="5443038326535759644",
                    ),
                ],
                [
                    InlineKeyboardButton(
                        text=" Instructions",
                        callback_data=f"inst_{user_id}",
                        style="danger",
                        icon_custom_emoji_id="5222444124698853913",
                    ),
                    InlineKeyboardButton(
                        text=" Language",
                        callback_data=f"lang_{user_id}",
                        style="danger",
                        icon_custom_emoji_id="5447410659077661506",
                    ),
                ],
                [
                    InlineKeyboardButton(
                        text=" Export",
                        callback_data=f"exp_{user_id}",
                        style="success",
                        icon_custom_emoji_id="5445355530111437729",
                    )
                ],
            ]
        )
    else:
        return InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=" Мой кабинет",
                        style="success",
                        icon_custom_emoji_id="5282843764451195532",
                        web_app=WebAppInfo(url=mini_app_url),
                    )
                ],
                [
                    InlineKeyboardButton(
                        text=" Подписка",
                        callback_data=f"sub_{user_id}",
                        style="primary",
                        icon_custom_emoji_id="5438496463044752972",
                    ),
                    InlineKeyboardButton(
                        text=" Поддержка",
                        callback_data=f"sup_{user_id}",
                        style="primary",
                        icon_custom_emoji_id="5443038326535759644",
                    ),
                ],
                [
                    InlineKeyboardButton(
                        text=" Инструкцыя",
                        callback_data=f"inst_{user_id}",
                        style="danger",
                        icon_custom_emoji_id="5222444124698853913",
                    ),
                    InlineKeyboardButton(
                        text=" Язык/Language",
                        callback_data=f"lang_{user_id}",
                        style="danger",
                        icon_custom_emoji_id="5447410659077661506",
                    ),
                ],
                [
                    InlineKeyboardButton(
                        text=" Экспорт",
                        callback_data=f"exp_{user_id}",
                        style="success",
                        icon_custom_emoji_id="5445355530111437729",
                    )
                ],
            ]
        )


BACK_EMOJI_ID = "5352759161945867747"


def back_keyboard(lang="ru"):
    """Keyboard with only a Back button"""
    label = " Назад" if lang == "ru" else " Back"
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=label,
                    callback_data="back_to_menu",
                    icon_custom_emoji_id=BACK_EMOJI_ID,
                )
            ]
        ]
    )


def sub_keyboard(lang="ru"):
    """Subscription keyboard: Buy + Back"""
    buy_label = " Купить подписку" if lang == "ru" else " Buy Subscription"
    back_label = " Назад" if lang == "ru" else " Back"
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=buy_label,
                    callback_data="buy_sub",
                    style="success",
                    icon_custom_emoji_id="5440841102871517055",
                )
            ],
            [
                InlineKeyboardButton(
                    text=back_label,
                    callback_data="back_to_menu",
                    icon_custom_emoji_id=BACK_EMOJI_ID,
                )
            ],
        ]
    )


def plans_keyboard(lang="ru", user_id=None):
    """Subscription plan selection keyboard"""
    back_label = " Назад" if lang == "ru" else " Back"
    back_callback = f"sub_{user_id}" if user_id else "back_to_menu"
    if lang == "ru":
        plans = [
            (" 1 месяц", "buy_plan_sub_1m", "5221507910702818314"),
            (" 3 месяца (-17%)", "buy_plan_sub_3m", "5229064374403998351"),
            (" 6 месяцев (-25%)", "buy_plan_sub_6m", "5229064374403998351"),
            (" 1 год (-42%)", "buy_plan_sub_12m", "5406683434124859552"),
        ]
    else:
        plans = [
            (" 1 month", "buy_plan_sub_1m", "5221507910702818314"),
            (" 3 months (-17%)", "buy_plan_sub_3m", "5229064374403998351"),
            (" 6 months (-25%)", "buy_plan_sub_6m", "5229064374403998351"),
            (" 1 year (-42%)", "buy_plan_sub_12m", "5406683434124859552"),
        ]
    rows = [
        [
            InlineKeyboardButton(
                text=lbl, callback_data=cb, icon_custom_emoji_id=emoji_id
            )
        ]
        for lbl, cb, emoji_id in plans
    ]
    rows.append(
        [
            InlineKeyboardButton(
                text=back_label,
                callback_data=back_callback,
                icon_custom_emoji_id=BACK_EMOJI_ID,
            )
        ]
    )
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.message(Command("start"))
async def cmd_start(message: types.Message, command: CommandObject):
    """Обработчик команды /start"""
    user_id = message.from_user.id
    user_name = message.from_user.first_name
    user_full_name = message.from_user.full_name
    start_args = command.args or ""

    # Сохраняем имя и username пользователя
    save_user_info(user_id, user_full_name, message.from_user.username)

    # Проверка блокировки
    if is_user_blocked(user_id):
        _blang = get_user_language(user_id) or "ru"
        await message.answer(
            "🚫 Your account has been blocked. Please contact the administrator."
            if _blang == "en"
            else "🚫 Ваш аккаунт заблокирован. Обратитесь к администратору."
        )
        return

    # Проверка подписки на канал
    is_subscribed = await check_channel_subscription(user_id)

    if not is_subscribed:
        # Создаем кнопки для подписки на каналы и проверки
        keyboard_buttons = []
        for channel_name, channel_username in REQUIRED_CHANNELS.items():
            keyboard_buttons.append(
                InlineKeyboardButton(
                    text=f" {channel_name}",
                    url=f"https://t.me/{channel_username.replace('@', '')}",
                    icon_custom_emoji_id="5298609163465692994",
                )
            )
        _clang = get_user_language(user_id) or "ru"
        keyboard_buttons.append(
            InlineKeyboardButton(
                text="Check subscription" if _clang == "en" else "Проверить подписку",
                callback_data=f"check_sub_{user_id}",
                icon_custom_emoji_id="5260463209562776385",
            )
        )

        # Создаем клавиатуру с кнопками (по 1 в ряд)
        keyboard_rows = [[btn] for btn in keyboard_buttons]
        check_sub_keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_rows)

        # Формируем список каналов
        channels_list = "\n".join(
            [f"• {name} {username}" for name, username in REQUIRED_CHANNELS.items()]
        )

        if _clang == "en":
            await message.answer(
                f'<b><tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> To use the bot you need to subscribe to the channels:</b>\n\n'
                '<tg-emoji emoji-id="5260612807568668744">👇</tg-emoji> After subscribing, press the button below to verify.',
                parse_mode="HTML",
                reply_markup=check_sub_keyboard,
            )
        else:
            await message.answer(
                f'<b><tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> Для использования бота необходимо подписаться на каналы:</b>\n\n'
                '<tg-emoji emoji-id="5260612807568668744">👇</tg-emoji> После подписки нажмите кнопку ниже для проверки.',
                parse_mode="HTML",
                reply_markup=check_sub_keyboard,
            )
        return

    # Проверяем, выбран ли язык
    user_language = get_user_language(user_id)

    # Если язык не выбран (новый пользователь), показываем выбор языка
    if not user_language:
        # Создаем клавиатуру с кнопками выбора языка
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=" Русский",
                        callback_data="ru",
                        icon_custom_emoji_id="5398017006165305287",
                    ),
                    InlineKeyboardButton(
                        text=" English",
                        callback_data="en",
                        icon_custom_emoji_id="5474446335744680905",
                    ),
                ]
            ]
        )

        # Текст приветствия с выбором языка
        welcome_text = f'<b><tg-emoji emoji-id="5343984088493599366">👋</tg-emoji> Привет/Hello, {user_full_name}!</b>\n\n<tg-emoji emoji-id="5260284397189346108">🔴</tg-emoji> Выбери свой язык:\n<tg-emoji emoji-id="5260284397189346108">🔴</tg-emoji> Choose your language:'

        # Отправляем изображение с приветствием
        try:
            await message.answer_photo(
                photo=FSInputFile(LANGUAGE_IMAGE_PATH),
                caption=welcome_text,
                reply_markup=keyboard,
                parse_mode="HTML",
            )
        except Exception:
            await message.answer(
                welcome_text,
                reply_markup=keyboard,
                parse_mode="HTML",
            )
        return

    # Если пользователь открыл бот через deep link ?start=export — показываем раздел экспорта
    if start_args == "export":
        if user_language == "en":
            await message.answer(
                "📤 <b>Data Export</b>\n\n"
                'You can export your data from the "My Cabinet" section.\n\n'
                "Go to the cabinet and tap the export button.",
                parse_mode="HTML",
                reply_markup=back_keyboard("en"),
            )
        else:
            await message.answer(
                '<tg-emoji emoji-id="5395755469660762251">📤</tg-emoji> <b>Выберите период для выгрузки данных в <u>.XLSX</u></b>',
                parse_mode="HTML",
                reply_markup=back_keyboard("ru"),
            )
        return

    # Создаем главное меню клавиатуры
    mini_app_url_with_user = f"{MINI_APP_URL}?user_id={user_id}"

    main_menu_keyboard = build_main_menu_keyboard(
        user_id, mini_app_url_with_user, user_language
    )

    if user_language == "ru":
        ru_caption = (
            '<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>С возвращением!</b>\n'
            '<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Добро пожаловать в лучшего бота для контроля ваших финансов и бизнеса!\n\n'
            '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Выберите нужный раздел из меню ниже:\n'
        )
        # Отправляем изображение с приветствием на русском
        try:
            await message.answer_photo(
                photo=FSInputFile(START_IMAGE_PATH),
                caption=ru_caption,
                reply_markup=main_menu_keyboard,
                parse_mode="HTML",
            )
        except Exception:
            await message.answer(
                ru_caption,
                reply_markup=main_menu_keyboard,
                parse_mode="HTML",
            )
    else:
        en_caption = (
            '<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>Welcome back!</b>\n'
            '<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Welcome to the best bot for managing your finances and business!\n\n'
            '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Select the desired section from the menu below:\n'
        )
        # Отправляем изображение с приветствием на английском
        try:
            await message.answer_photo(
                photo=FSInputFile(START_IMAGE_PATH),
                caption=en_caption,
                reply_markup=main_menu_keyboard,
                parse_mode="HTML",
            )
        except Exception:
            await message.answer(
                en_caption,
                reply_markup=main_menu_keyboard,
                parse_mode="HTML",
            )


@router.callback_query(F.data == "expcustom")
async def handle_export_custom(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик кнопки 'Указать произвольные даты'"""
    _lang = get_user_language(callback.from_user.id) or "ru"
    if _lang == "en":
        await callback.message.answer(
            '<tg-emoji emoji-id="5413879192267805083">📅</tg-emoji> <b>Custom date range</b>\n\n'
            "Enter the <b>start date</b> in the format <code>DD.MM.YYYY</code>\n",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="Cancel",
                            callback_data="expcancel",
                            icon_custom_emoji_id="5807626765874499116",
                        )
                    ]
                ]
            ),
        )
    else:
        await callback.message.answer(
            '<tg-emoji emoji-id="5413879192267805083">📅</tg-emoji> <b>Произвольный диапазон дат</b>\n\n'
            "Введите <b>начальную дату</b> в формате <code>ДД.ММ.ГГГГ</code>\n",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="Отменить",
                            callback_data="expcancel",
                            icon_custom_emoji_id="5807626765874499116",
                        )
                    ]
                ]
            ),
        )
    await state.set_state(ExportState.waiting_start_date)
    await callback.answer()


@router.callback_query(F.data == "expcancel")
async def handle_export_cancel(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик отмены экспорта"""
    await state.clear()
    user_id = callback.from_user.id
    _lang = get_user_language(user_id) or "ru"
    months = get_months_with_operations(user_id)
    if months:
        if _lang == "en":
            await callback.message.answer(
                '<tg-emoji emoji-id="5445355530111437729">📤</tg-emoji> <b>Select a period to export data to <u>.XLSX</u></b>\n\n'
                "Choose a month or specify a custom date range:",
                parse_mode="HTML",
                reply_markup=build_export_months_keyboard(user_id, _lang),
            )
        else:
            await callback.message.answer(
                '<tg-emoji emoji-id="5445355530111437729">📤</tg-emoji> <b>Выберите период для выгрузки данных в <u>.XLSX</u></b>\n\n'
                "Выберите месяц или укажите произвольный диапазон дат:",
                parse_mode="HTML",
                reply_markup=build_export_months_keyboard(user_id, _lang),
            )
    else:
        _back_label = " Back" if _lang == "en" else " Назад"
        _cancel_text = "Export cancelled." if _lang == "en" else "Экспорт отменён."
        await callback.message.answer(
            _cancel_text,
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text=_back_label,
                            callback_data="back_to_menu",
                            icon_custom_emoji_id="5352759161945867747",
                        )
                    ]
                ]
            ),
        )
    await callback.answer()


@router.callback_query()
async def handle_callbacks(callback: types.CallbackQuery):
    """Обработчик callback запросов"""

    # Получаем ID пользователя
    user_id = callback.from_user.id

    # Проверка блокировки
    if is_user_blocked(user_id):
        blocked_lang = get_user_language(user_id) or "ru"
        blocked_msg = (
            "🚫 Your account has been blocked."
            if blocked_lang == "en"
            else "🚫 Ваш аккаунт заблокирован."
        )
        await callback.answer(blocked_msg, show_alert=True)
        return

    # Удаляем сообщение с клавиатурой выбора языка
    await callback.message.delete()

    # Создаем URL с ID пользователя
    mini_app_url_with_user = f"{MINI_APP_URL}?user_id={user_id}"

    # Определяем язык пользователя (до установки, чтобы использовать в check_sub_ и т.д.)
    current_lang = get_user_language(user_id) or "ru"

    # Клавиатура будет строиться после установки языка (для ru/en колбэков она перестроится)
    main_menu_keyboard = build_main_menu_keyboard(
        user_id, mini_app_url_with_user, current_lang
    )

    # Обработка проверки подписки
    if callback.data.startswith("check_sub_"):
        is_subscribed = await check_channel_subscription(user_id)
        if is_subscribed:
            # Проверяем, выбран ли язык
            user_language = get_user_language(user_id)

            if user_language == "ru":
                await callback.message.answer_photo(
                    photo=FSInputFile(START_IMAGE_PATH),
                    caption='<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>Отлично! Вы подписаны!</b>\n'
                    '<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Добро пожаловать в лучшего бота для контроля ваших финансов и бизнеса!\n\n'
                    '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Выберите нужный раздел из меню ниже:\n',
                    reply_markup=main_menu_keyboard,
                    parse_mode="HTML",
                )
            else:
                await callback.message.answer_photo(
                    photo=FSInputFile(START_IMAGE_PATH),
                    caption='<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>Great! You are subscribed!</b>\n'
                    '<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Welcome to the best bot for managing your finances and business!\n\n'
                    '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Select the desired section from the menu below:\n',
                    reply_markup=main_menu_keyboard,
                    parse_mode="HTML",
                )
        else:
            # Создаем кнопки для подписки на каналы и проверки
            keyboard_buttons = []
            for channel_name, channel_username in REQUIRED_CHANNELS.items():
                keyboard_buttons.append(
                    InlineKeyboardButton(
                        text=f'<tg-emoji emoji-id="5424818078833715060">📢</tg-emoji> {channel_name}',
                        url=f"https://t.me/{channel_username.replace('@', '')}",
                    )
                )
            _check_btn_text = (
                "✅ Check subscription"
                if current_lang == "en"
                else '<tg-emoji emoji-id="5260463209562776385">✅</tg-emoji> Проверить подписку'
            )
            keyboard_buttons.append(
                InlineKeyboardButton(
                    text=_check_btn_text,
                    callback_data=f"check_sub_{user_id}",
                )
            )

            # Создаем клавиатуру с кнопками (по 1 в ряд)
            keyboard_rows = [[btn] for btn in keyboard_buttons]
            check_sub_keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_rows)

            # Формируем список каналов
            channels_list = "\n".join(
                [f"• {name} {username}" for name, username in REQUIRED_CHANNELS.items()]
            )

            if current_lang == "en":
                await callback.message.answer(
                    f"<b>⚠️ You haven't subscribed to the channels yet:</b>\n\n{channels_list}\n\n"
                    "Please subscribe and press the button below to verify.",
                    parse_mode="HTML",
                    reply_markup=check_sub_keyboard,
                )
            else:
                await callback.message.answer(
                    f"<b>⚠️ Вы ещё не подписались на каналы:</b>\n\n{channels_list}\n\n"
                    "Пожалуйста, подпишитесь и нажмите кнопку ниже для проверки.",
                    parse_mode="HTML",
                    reply_markup=check_sub_keyboard,
                )
        await callback.answer()
        return

    # Сохраняем язык при первом выборе (ru или en)
    if callback.data == "ru":
        set_user_language(user_id, "ru")
        init_trial_subscription(user_id)
        ru_keyboard = build_main_menu_keyboard(user_id, mini_app_url_with_user, "ru")
        await callback.message.answer_photo(
            photo=FSInputFile(START_IMAGE_PATH),
            caption='<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>Отлично!</b>\n<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Добро пожаловать в лучшего бота для контроля ваших финансов и бизнеса!\n<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Вам доступен <b>7-дневный бесплатный пробный период</b>, чтобы попробовать все возможности бота.\n\n'
            '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Выберите нужный раздел из меню ниже:\n\n',
            reply_markup=ru_keyboard,
            parse_mode="HTML",
        )
    elif callback.data == "en":
        set_user_language(user_id, "en")
        init_trial_subscription(user_id)
        en_keyboard = build_main_menu_keyboard(user_id, mini_app_url_with_user, "en")
        await callback.message.answer_photo(
            photo=FSInputFile(START_IMAGE_PATH),
            caption='<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>Great!</b>\n<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Welcome to the best bot for managing your finances and business!\n<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> You have access to a <b>7-day free trial period</b> so you can explore and test all the features of the bot.\n\n'
            '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Select the desired section from the menu below:\n',
            reply_markup=en_keyboard,
            parse_mode="HTML",
        )

    # Назад — возврат в главное меню
    elif callback.data == "back_to_menu":
        if current_lang == "ru":
            caption = (
                '<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>С возвращением!</b>\n'
                '<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Добро пожаловать в лучшего бота для контроля ваших финансов и бизнеса!\n\n'
                '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Выберите нужный раздел из меню ниже:\n'
            )
        else:
            caption = (
                '<tg-emoji emoji-id="5395755469660762251">🎊</tg-emoji> <b>Welcome back!</b>\n'
                '<tg-emoji emoji-id="5258354775757439405">👉</tg-emoji> Welcome to the best bot for managing your finances and business!\n\n'
                '<tg-emoji emoji-id="5231102735817918643">👇</tg-emoji> Select the desired section from the menu below:\n'
            )
        try:
            await callback.message.answer_photo(
                photo=FSInputFile(START_IMAGE_PATH),
                caption=caption,
                reply_markup=main_menu_keyboard,
                parse_mode="HTML",
            )
        except Exception:
            await callback.message.answer(
                caption, reply_markup=main_menu_keyboard, parse_mode="HTML"
            )

    # Показать выбор плана подписки
    elif callback.data == "buy_sub":
        if current_lang == "ru":
            await callback.message.answer(
                '<tg-emoji emoji-id="5438496463044752972">⭐️</tg-emoji> <b><u>Выберите срок доступа. Чем дольше период — тем дешевле выходит месяц.</u></b>\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>1 месяц — 30 дней</b>\n<b>199₽</b> или 109 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>3 месяца — 90 дней</b>\n<s>597₽</s> → <b>499₽</b> или 289 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n≈ 166₽ в месяц (<b>экономия 17%</b>)\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>6 месяцев — 180 дней</b>\n<s>1194₽</s> → <b>899₽</b> или 539 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n≈ 149₽ в месяц (<b>экономия 25%</b>)\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>1 год — 365 дней</b> <tg-emoji emoji-id="5424972470023104089">🔥</tg-emoji> <b>Самый выгодный вариант</b>\n<s>2421₽</s> → <b>1399₽</b> или 849 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n≈ 115₽ в месяц (<b>экономия 42% — максимальная скидка</b>)\n\n'
                '<tg-emoji emoji-id="5445353829304387411">💳</tg-emoji> Оплатить проходит пока что только через <b>Telegram Stars.</b>\n\n'
                "❌ <b>Подписка не продлевается автоматически.</b>\n\n"
                '<tg-emoji emoji-id="5208573502046610594">🏪</tg-emoji> Продлить доступ можно в любое время — оставшиеся дни сохраняются и добавляются к новому сроку.',
                parse_mode="HTML",
                reply_markup=plans_keyboard("ru", user_id),
            )
        else:
            await callback.message.answer(
                '<tg-emoji emoji-id="5438496463044752972">⭐️</tg-emoji> <b><u>Choose your access period. The longer the period — the cheaper each month.</u></b>\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>1 month — 30 days</b>\n<b>199₽</b> or 109 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>3 months — 90 days</b>\n<s>597₽</s> → <b>499₽</b> or 289 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n≈ 166₽ per month (<b>save 17%</b>)\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>6 months — 180 days</b>\n<s>1194₽</s> → <b>899₽</b> or 539 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n≈ 149₽ per month (<b>save 25%</b>)\n\n'
                '<tg-emoji emoji-id="5258354775757439405">▶️</tg-emoji> <b>1 year — 365 days</b> <tg-emoji emoji-id="5424972470023104089">🔥</tg-emoji> <b>Best value</b>\n<s>2421₽</s> → <b>1399₽</b> or 849 <tg-emoji emoji-id="5321485469249198987">⭐️</tg-emoji>\n≈ 115₽ per month (<b>save 42% — maximum discount</b>)\n\n'
                '<tg-emoji emoji-id="5445353829304387411">💳</tg-emoji> Payment is currently available only via <b>Telegram Stars.</b>\n\n'
                "❌ <b>Subscription does not renew automatically.</b>\n\n"
                '<tg-emoji emoji-id="5208573502046610594">🏪</tg-emoji> You can renew access at any time — remaining days are saved and added to the new period.',
                parse_mode="HTML",
                reply_markup=plans_keyboard("en", user_id),
            )

    # Покупка конкретного плана — отправляем инвойс
    elif callback.data.startswith("buy_plan_"):
        plan_map = {
            "buy_plan_sub_1m": ("NytroBC — 1 месяц", "sub_1m", 109),
            "buy_plan_sub_3m": ("NytroBC — 3 месяца", "sub_3m", 289),
            "buy_plan_sub_6m": ("NytroBC — 6 месяцев", "sub_6m", 539),
            "buy_plan_sub_12m": ("NytroBC — 1 год", "sub_12m", 849),
        }
        plan_info = plan_map.get(callback.data)
        if plan_info:
            title, payload, stars = plan_info
            await bot.send_invoice(
                chat_id=user_id,
                title=title,
                description="Полный доступ ко всем функциям NytroBC",
                payload=payload,
                provider_token="",
                currency="XTR",
                prices=[LabeledPrice(label=title, amount=stars)],
            )

    # Обработка callback от кнопок меню
    elif callback.data.startswith("sub_"):
        init_trial_subscription(user_id)
        sub_info = get_subscription_info(user_id)
        sub_active = sub_info["active"]
        sub_end_date = sub_info["subscription_end"]
        if current_lang == "en":
            status_icon = "🟢" if sub_active else "🔴"
            status_text = "Active" if sub_active else "Expired"
            end_text = sub_end_date if sub_end_date else "N/A"
            await callback.message.answer(
                '<tg-emoji emoji-id="5377599075237502153">💳</tg-emoji> <b>Subscription</b>\n\n'
                '<tg-emoji emoji-id="5427168083074628963">💎</tg-emoji> With an active subscription you get <b>full control over your finances and all bot features:</b>\n\n'
                "• Adding and editing <b>transactions</b>\n"
                "• Managing <b>categories</b> and <b>profile</b>\n"
                "• Tracking <b>savings</b> and <b>budget planning</b>\n"
                "• Access to <b>instructions</b> and useful <b>materials</b>\n"
                "• Using the <b>Mini App</b> with a <b>beautiful</b> and <b>intuitive</b> interface\n"
                "• Creating <b>plans</b> and <b>notifications</b> with reminders\n"
                "• Viewing full <b>statistics</b> on your profile and <b>transactions</b>\n"
                "• <b>Data export</b> to <b>CSV</b>\n"
                "• A <b>star</b> in the Mini App before your name for <b>easy</b> identification\n"
                "• <b>Friendly</b> and <b>positive</b> support 24/7\n\n"
                '<tg-emoji emoji-id="5273914604752216432">❌</tg-emoji> Without a subscription editing and adding data is unavailable, but viewing all <b>statistics</b> and transaction history remains accessible. All your <b>data</b> is saved and will never disappear.\n\n'
                f"{status_icon} Status: <b>{status_text}</b>\n"
                f'<tg-emoji emoji-id="5413879192267805083">🗓</tg-emoji> Expiry date: <b>{end_text}</b>\n\n'
                '<tg-emoji emoji-id="5877458226823302157">🕒</tg-emoji> Subscription validity is calculated in Moscow time (UTC+3).',
                parse_mode="HTML",
                reply_markup=sub_keyboard("en"),
            )
        else:
            status_icon = "🟢" if sub_active else "🔴"
            status_text = "Активный" if sub_active else "Истекший"
            end_text = sub_end_date if sub_end_date else "НЕТ"
            await callback.message.answer(
                '<tg-emoji emoji-id="5377599075237502153">💳</tg-emoji> <b>Подписка</b>\n\n'
                '<tg-emoji emoji-id="5427168083074628963">💎</tg-emoji> С активной подпиской ты получаешь <b>полный контроль над своими финансами и всеми функциями бота:</b>\n\n'
                "• Добавление и редактирование <b>операций</b>\n"
                "• Настройка <b>категорий</b> и <b>профиля</b>\n"
                "• Контроль <b>накоплений</b> и <b>планирование бюджета</b>\n"
                "• Доступ к <b>инструкциям</b> и полезным <b>материалам</b>\n"
                "• Использование <b>Mini App</b> с <b>красивым</b> и <b>понятным</b> интерфейсом\n"
                "• Создание <b>планов</b> и <b>уведомлений</b> с напоминаниями\n"
                "• Просмотр полной <b>статистики</b> по профилю и <b>операциям</b>\n"
                "• <b>Экспорт данных</b> в <b>CSV</b>\n"
                "• <b>Звезда</b> в Mini App перед именем для <b>удобной</b> идентификации\n"
                "• <b>Добрая</b> и <b>позитивная</b> поддержка 24/7\n\n"
                '<tg-emoji emoji-id="5273914604752216432">❌</tg-emoji> Без подписки редактирование и добавление данных недоступно, но просмотр всей <b>статистики</b> и истории <b>операций</b> остаётся доступным. Все твои <b>данные</b> сохраняются и никуда не исчезнут.\n\n'
                f"{status_icon} Статус: <b>{status_text}</b>\n"
                f'<tg-emoji emoji-id="5413879192267805083">🗓</tg-emoji> Дата окончания: <b>{end_text}</b>\n\n'
                '<tg-emoji emoji-id="5877458226823302157">🕒</tg-emoji> Срок действия подписки рассчитывается по московскому времени (UTC+3).',
                parse_mode="HTML",
                reply_markup=sub_keyboard("ru"),
            )
    elif callback.data.startswith("sup_"):
        if current_lang == "en":
            await callback.message.answer(
                '<tg-emoji emoji-id="5443038326535759644">💬</tg-emoji> <b>Support</b>\n\n'
                '<tg-emoji emoji-id="5436113877181941026">❓</tg-emoji> For all questions and deals, contact: @s_narzimurodov\n\n'
                '<tg-emoji emoji-id="5208573502046610594">🏪</tg-emoji> We are ready to help you 24/7!',
                parse_mode="HTML",
                reply_markup=back_keyboard("en"),
            )
        else:
            await callback.message.answer(
                '<tg-emoji emoji-id="5443038326535759644">💬</tg-emoji> <b>Служба поддержки</b>\n\n'
                '<tg-emoji emoji-id="5436113877181941026">❓</tg-emoji> По всем вопросам и сделкам обращайтесь: @s_narzimurodov\n\n'
                '<tg-emoji emoji-id="5208573502046610594">🏪</tg-emoji> Мы готовы помочь вам 24/7!',
                parse_mode="HTML",
                reply_markup=back_keyboard("ru"),
            )
    elif callback.data.startswith("inst_"):
        if current_lang == "en":
            await callback.message.answer(
                """<tg-emoji emoji-id="5226512880362332956">📖</tg-emoji> NytroBC Bot User Guide

                Welcome to NytroBC — your personal financial assistant!


                <tg-emoji emoji-id="5310278924616356636">🎯</tg-emoji> Main Features

                1. <b>My Cabinet</b>  
                Open your personal dashboard to manage your finances. Here you can:  
                - Add income and expenses  
                - Create your own categories  
                - Set payment reminders  
                - Plan your budget  
                - Track your savings  

                2. <b>Subscription</b>  
                Full access to the bot requires a subscription. On your first login, you get a <b>7-day free trial.</b>  

                3. <b>Support</b>  
                Contact the support team if you have any questions or issues.  

                4. <b>Language</b>  
                You can switch the interface language:  
                - <tg-emoji emoji-id="5449408995691341691">🇷🇺</tg-emoji> Russian  
                - <tg-emoji emoji-id="5202196682497859879">🇬🇧</tg-emoji> English  

                5. <b>Data Export</b>  
                Export your financial data in Excel (XLSX) format for:  
                - A specific month  
                - A custom period (select dates manually)  


                <tg-emoji emoji-id="5188481279963715781">🚀</tg-emoji> Getting Started

                Step 1: Launch the bot  
                Press the <b>/start</b> button in the chat with the bot.  

                Step 2: Subscribe to the channel  
                To use the bot, you must be subscribed to <b>@nbc_channeI</b>. The bot will automatically check your subscription.  

                Step 3: Open My Cabinet  
                Click the <b>“My Cabinet”</b> button to open the web app and start managing your finances.  


                Tips for using the bot

                1. <b>Track expenses regularly</b> — get a clear picture of your finances  
                2. <b>Use categories</b> — organize income and expenses for better analysis  
                3. <b>Set reminders</b> — don’t miss important payments  
                4. <b>Export your data</b> — create backups and analyze your stats  


                <tg-emoji emoji-id="5436113877181941026">❓</tg-emoji> Frequently Asked Questions

                <b>How do I get the free trial?</b>  
                > The trial subscription is activated automatically on your first launch and lasts 7 days.  

                <b>How do I extend my subscription?</b>  
                > Click the “Subscription” button in the bot menu and choose a plan.  

                <b>How do I change the language?</b>  
                > Click “Language” in the menu and select your preferred option.  


                Enjoy using NytroBC! <tg-emoji emoji-id="5395755469660762251">🎉</tg-emoji>""",
                parse_mode="HTML",
                reply_markup=back_keyboard("en"),
            )
        else:
            await callback.message.answer(
                """<tg-emoji emoji-id="5226512880362332956">📖</tg-emoji> Инструкция по использованию бота NytroBC

Добро пожаловать в NytroBC — ваш личный финансовый помощник!


 <tg-emoji emoji-id="5310278924616356636">🎯</tg-emoji> Основные возможности бота

1. <b>Мой кабинет</b>
Откройте личный кабинет для управления финансами. Здесь вы можете:
- Добавлять доходы и расходы
- Создавать собственные категории
- Устанавливать напоминания о платежах
- Планировать бюджет
- Отслеживать накопления

2. <b>Подписка</b>
Для полного доступа к функционалу бота требуется подписка. При первом входе предоставляется <b>бесплатная пробная подписка на 7 дней.</b>

3. <b>Поддержка</b>
Свяжитесь с командой поддержки для решения любых вопросов.

4. <b>Язык / Language</b>
Вы можете переключить язык интерфейса:
-  <tg-emoji emoji-id="5449408995691341691">🇷🇺</tg-emoji> Русский
-  <tg-emoji emoji-id="5202196682497859879">🇬🇧</tg-emoji> English

5. <b>Экспорт данных</b>
Экспортируйте свои финансовые данные в формате Excel (XLSX) за:
- Конкретный месяц
- Произвольный период (указать даты вручную)


 <tg-emoji emoji-id="5188481279963715781">🚀</tg-emoji> Как начать пользоваться

Шаг 1: Запуск бота
Нажмите кнопку <b>/start</b> в чате с ботом.

Шаг 2: Подписка на канал
Для использования бота необходимо быть подписанным на канал <b>@nbc_channeI</b>. Бот автоматически проверит вашу подписку.

Шаг 3: Откройте Мой кабинет
Нажмите кнопку <b>«Мой кабинет»</b> (или My Cabinet), чтобы открыть веб-приложение для управления финансами.


 Советы по использованию

1. <b>Регулярно записывайте расходы</b> — так вы получите точную картину своих финансов
2. <b>Используйте категории</b> — группируйте доходы и расходы для удобного анализа
3. <b>Настройте напоминания</b> — не пропускайте важные платежи
4. <b>Экспортируйте данные</b> — создавайте резервные копии и анализируйте статистику


 <tg-emoji emoji-id="5436113877181941026">❓</tg-emoji> Частые вопросы

<b>Как получить пробную подписку?</b>
> При первом запуске бота подписка активируется автоматически на 7 дней.

<b>Как продлить подписку?</b>
> Нажмите кнопку «Подписка» в меню бота и выберите удобный тариф.

<b>Как изменить язык?</b>
> Нажмите кнопку «Язык/Language» в меню и выберите нужный язык.


Приятного использования!  <tg-emoji emoji-id="5395755469660762251">🎉</tg-emoji>""",
                parse_mode="HTML",
                reply_markup=back_keyboard("ru"),
            )
    elif callback.data.startswith("lang_"):
        language_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=" Русский",
                        callback_data="ru",
                        icon_custom_emoji_id="5398017006165305287",
                    ),
                    InlineKeyboardButton(
                        text=" English",
                        callback_data="en",
                        icon_custom_emoji_id="5474446335744680905",
                    ),
                ]
            ]
        )
        await callback.message.answer(
            '<tg-emoji emoji-id="5447410659077661506">🌐</tg-emoji> Выберите язык:\nChoose your language:',
            reply_markup=language_keyboard,
            parse_mode="HTML",
        )
    elif callback.data.startswith("exp_"):
        months = get_months_with_operations(user_id)
        if not months:
            if current_lang == "en":
                await callback.message.answer(
                    '<tg-emoji emoji-id="5445355530111437729">📤</tg-emoji> <b>Data Export</b>\n\nYou have no operations to export yet.',
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text=" Back",
                                    callback_data="back_to_menu",
                                    icon_custom_emoji_id="5352759161945867747",
                                )
                            ]
                        ]
                    ),
                )
            else:
                await callback.message.answer(
                    '<tg-emoji emoji-id="5445355530111437729">📤</tg-emoji> <b>Экспорт данных</b>\n\nУ вас пока нет операций для экспорта.',
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text=" Назад",
                                    callback_data="back_to_menu",
                                    icon_custom_emoji_id="5352759161945867747",
                                )
                            ]
                        ]
                    ),
                )
        else:
            if current_lang == "en":
                await callback.message.answer(
                    '<tg-emoji emoji-id="5445355530111437729">📤</tg-emoji> <b>Select a period to export data to <u>.XLSX</u></b>\n\n'
                    "Choose a month or specify a custom date range:",
                    parse_mode="HTML",
                    reply_markup=build_export_months_keyboard(user_id, "en"),
                )
            else:
                await callback.message.answer(
                    '<tg-emoji emoji-id="5445355530111437729">📤</tg-emoji> <b>Выберите период для выгрузки данных в <u>.XLSX</u></b>\n\n'
                    "Выберите месяц или укажите произвольный диапазон дат:",
                    parse_mode="HTML",
                    reply_markup=build_export_months_keyboard(user_id, "ru"),
                )

    elif callback.data.startswith("expm_"):
        parts = callback.data.split("_")
        year, month = int(parts[1]), int(parts[2])
        start_date = datetime(year, month, 1).date()
        end_day = days_in_month(year, month)
        end_date = datetime(year, month, end_day).date()
        start_str = start_date.strftime("%Y%m%d")
        end_str = end_date.strftime("%Y%m%d")
        if current_lang == "en":
            start_label = format_date_en(year, month, 1)
            end_label = format_date_en(year, month, end_day)
            await callback.message.answer(
                f'<tg-emoji emoji-id="5904219717073114606">📤</tg-emoji> <b>Export confirmation</b>\n\n'
                f"Period: <b>{start_label} — {end_label}</b>\n\n"
                "Press «Export» to receive the .XLSX file:",
                parse_mode="HTML",
                reply_markup=build_export_confirm_keyboard(start_str, end_str, "en"),
            )
        else:
            start_label = format_date_ru(year, month, 1)
            end_label = format_date_ru(year, month, end_day)
            await callback.message.answer(
                f'<tg-emoji emoji-id="5904219717073114606">📤</tg-emoji> <b>Подтверждение экспорта</b>\n\n'
                f"Период: <b>{start_label} — {end_label}</b>\n\n"
                "Нажмите «Экспортировать» для получения файла .XLSX:",
                parse_mode="HTML",
                reply_markup=build_export_confirm_keyboard(start_str, end_str, "ru"),
            )

    elif callback.data == "expcustom":
        pass

    elif callback.data.startswith("expdo_"):
        exp_sub_info = get_subscription_info(user_id)
        if not exp_sub_info["active"]:
            if current_lang == "en":
                await callback.message.answer(
                    '<tg-emoji emoji-id="5370553005294754329">😩</tg-emoji> Unfortunately your subscription is not active! You can purchase it by pressing the button below',
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text=" Buy Subscription",
                                    callback_data=f"sub_{user_id}",
                                    icon_custom_emoji_id="5287231198098117669",
                                    style="success",
                                )
                            ]
                        ]
                    ),
                )
            else:
                await callback.message.answer(
                    '<tg-emoji emoji-id="5370553005294754329">😩</tg-emoji> К сожалению у вас не активна подписка! Вы можете приобрести её нажав на кнопку ниже',
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text=" Купить подписку",
                                    callback_data=f"sub_{user_id}",
                                    icon_custom_emoji_id="5287231198098117669",
                                    style="success",
                                )
                            ]
                        ]
                    ),
                )
            await callback.answer()
            return
        parts = callback.data.split("_")
        start_str = parts[1]
        end_str = parts[2]
        start_date = datetime.strptime(start_str, "%Y%m%d").date()
        end_date = datetime.strptime(end_str, "%Y%m%d").date()
        records = get_finance_in_range(user_id, start_date, end_date)
        if not records:
            if current_lang == "en":
                await callback.message.answer(
                    '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> No operations found for the specified period.',
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text=" Back",
                                    callback_data=f"exp_{user_id}",
                                    icon_custom_emoji_id="5352759161945867747",
                                )
                            ]
                        ]
                    ),
                )
            else:
                await callback.message.answer(
                    '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> За указанный период операций не найдено.',
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text=" Назад",
                                    callback_data=f"exp_{user_id}",
                                    icon_custom_emoji_id="5352759161945867747",
                                )
                            ]
                        ]
                    ),
                )
        else:
            xlsx_bytes = generate_export_xlsx(records)
            if current_lang == "en":
                s_label = format_date_en(
                    start_date.year, start_date.month, start_date.day
                )
                e_label = format_date_en(end_date.year, end_date.month, end_date.day)
                filename = f"export_{start_str}_{end_str}.xlsx"
                await callback.message.answer_document(
                    document=BufferedInputFile(xlsx_bytes, filename=filename),
                    caption=f'<tg-emoji emoji-id="5206607081334906820">✅</tg-emoji> Export completed\nPeriod: {s_label} — {e_label}\nOperations: {len(records)}',
                    parse_mode="HTML",
                )
            else:
                s_label = format_date_ru(
                    start_date.year, start_date.month, start_date.day
                )
                e_label = format_date_ru(end_date.year, end_date.month, end_date.day)
                filename = f"export_{start_str}_{end_str}.xlsx"
                await callback.message.answer_document(
                    document=BufferedInputFile(xlsx_bytes, filename=filename),
                    caption=f'<tg-emoji emoji-id="5206607081334906820">✅</tg-emoji> Экспорт завершён\nПериод: {s_label} — {e_label}\nОпераций: {len(records)}',
                    parse_mode="HTML",
                )

    elif callback.data == "expcancel":
        pass

    await callback.answer()


@router.pre_checkout_query()
async def pre_checkout_handler(pre_checkout_query: types.PreCheckoutQuery):
    """Подтверждение предварительного запроса оплаты"""
    await bot.answer_pre_checkout_query(pre_checkout_query.id, ok=True)


@router.message(F.successful_payment)
async def successful_payment_handler(message: types.Message):
    """Обработка успешной оплаты через Telegram Stars"""
    payload = message.successful_payment.invoice_payload
    user_id = message.from_user.id
    _lang = get_user_language(user_id) or "ru"

    plan_days = {"sub_1m": 30, "sub_3m": 90, "sub_6m": 180, "sub_12m": 365}
    days = plan_days.get(payload, 30)
    new_end = extend_subscription(user_id, days)

    if _lang == "en":
        await message.answer(
            f'<tg-emoji emoji-id="5438496463044752972">⭐️</tg-emoji> <b>Thank you for purchasing a subscription!</b>\n\n'
            f"Your subscription is activated until <b>{new_end.strftime('%d.%m.%Y')}</b>.\n"
            "Enjoy using NytroBC!",
            parse_mode="HTML",
        )
    else:
        await message.answer(
            f'<tg-emoji emoji-id="5438496463044752972">⭐️</tg-emoji> <b>Спасибо за приобретение подписки!</b>\n\n'
            f"Ваша подписка активирована до <b>{new_end.strftime('%d.%m.%Y')}</b>.\n"
            "Приятного использования NytroBC!",
            parse_mode="HTML",
        )


@router.message(ExportState.waiting_start_date)
async def handle_export_start_date(message: types.Message, state: FSMContext):
    """Обработчик ввода начальной даты для произвольного экспорта"""
    _lang = get_user_language(message.from_user.id) or "ru"
    date_str = message.text.strip() if message.text else ""
    d = parse_finance_date(date_str)
    if not d:
        if _lang == "en":
            await message.answer(
                '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> Invalid date format. Enter the date in the format <code>DD.MM.YYYY</code>\nFor example: <code>05.01.2026</code>',
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text="Cancel",
                                callback_data="expcancel",
                                icon_custom_emoji_id="5807626765874499116",
                            )
                        ]
                    ]
                ),
            )
        else:
            await message.answer(
                '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> Неверный формат даты. Введите дату в формате <code>ДД.ММ.ГГГГ</code>\nНапример: <code>05.01.2026</code>',
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text="Отменить",
                                callback_data="expcancel",
                                icon_custom_emoji_id="5807626765874499116",
                            )
                        ]
                    ]
                ),
            )
        return
    await state.update_data(start_date=d.strftime("%Y%m%d"))
    await state.set_state(ExportState.waiting_end_date)
    if _lang == "en":
        await message.answer(
            f'<tg-emoji emoji-id="5397916757333654639">✅</tg-emoji> Start date: <b>{format_date_en(d.year, d.month, d.day)}</b>\n\n'
            "Now enter the <b>end date</b> in the format <code>DD.MM.YYYY</code>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="Cancel",
                            callback_data="expcancel",
                            icon_custom_emoji_id="5807626765874499116",
                        )
                    ]
                ]
            ),
        )
    else:
        await message.answer(
            f'<tg-emoji emoji-id="5397916757333654639">✅</tg-emoji> Начальная дата: <b>{format_date_ru(d.year, d.month, d.day)}</b>\n\n'
            "Теперь введите <b>конечную дату</b> в формате <code>ДД.ММ.ГГГГ</code>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="Отменить",
                            callback_data="expcancel",
                            icon_custom_emoji_id="5807626765874499116",
                        )
                    ]
                ]
            ),
        )


@router.message(ExportState.waiting_end_date)
async def handle_export_end_date(message: types.Message, state: FSMContext):
    """Обработчик ввода конечной даты для произвольного экспорта"""
    _lang = get_user_language(message.from_user.id) or "ru"
    date_str = message.text.strip() if message.text else ""
    d = parse_finance_date(date_str)
    if not d:
        if _lang == "en":
            await message.answer(
                '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> Invalid date format. Enter the date in the format <code>DD.MM.YYYY</code>\nFor example: <code>10.01.2026</code>',
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text="Cancel",
                                callback_data="expcancel",
                                icon_custom_emoji_id="5807626765874499116",
                            )
                        ]
                    ]
                ),
            )
        else:
            await message.answer(
                '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> Неверный формат даты. Введите дату в формате <code>ДД.ММ.ГГГГ</code>\nНапример: <code>10.01.2026</code>',
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text="Отменить",
                                callback_data="expcancel",
                                icon_custom_emoji_id="5807626765874499116",
                            )
                        ]
                    ]
                ),
            )
        return
    data = await state.get_data()
    start_str = data.get("start_date")
    end_str = d.strftime("%Y%m%d")
    await state.clear()

    start_date = datetime.strptime(start_str, "%Y%m%d").date()
    end_date = d

    if end_date < start_date:
        if _lang == "en":
            await message.answer(
                '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> The end date cannot be earlier than the start date. Please try again.',
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text=" Back to export",
                                callback_data=f"exp_{message.from_user.id}",
                                icon_custom_emoji_id="5352759161945867747",
                            )
                        ]
                    ]
                ),
            )
        else:
            await message.answer(
                '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> Конечная дата не может быть раньше начальной. Попробуйте ещё раз.',
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text=" Назад к экспорту",
                                callback_data=f"exp_{message.from_user.id}",
                                icon_custom_emoji_id="5352759161945867747",
                            )
                        ]
                    ]
                ),
            )
        return

    if _lang == "en":
        start_label = format_date_en(start_date.year, start_date.month, start_date.day)
        end_label = format_date_en(end_date.year, end_date.month, end_date.day)
        await message.answer(
            f'<tg-emoji emoji-id="5904219717073114606">📤</tg-emoji> <b>Export confirmation</b>\n\n'
            f"Period: <b>{start_label} — {end_label}</b>\n\n"
            "Press «Export» to receive the .XLSX file:",
            parse_mode="HTML",
            reply_markup=build_export_confirm_keyboard(start_str, end_str, "en"),
        )
    else:
        start_label = format_date_ru(start_date.year, start_date.month, start_date.day)
        end_label = format_date_ru(end_date.year, end_date.month, end_date.day)
        await message.answer(
            f'<tg-emoji emoji-id="5904219717073114606">📤</tg-emoji> <b>Подтверждение экспорта</b>\n\n'
            f"Период: <b>{start_label} — {end_label}</b>\n\n"
            "Нажмите «Экспортировать» для получения файла .XLSX:",
            parse_mode="HTML",
            reply_markup=build_export_confirm_keyboard(start_str, end_str, "ru"),
        )


# Обработчики для кнопок главного меню
@router.message()
async def handle_menu_buttons(message: types.Message):
    """Обработчик нажатий на кнопки меню"""
    user_id = message.from_user.id
    user_full_name = message.from_user.full_name

    # Сохраняем имя и username пользователя
    save_user_info(user_id, user_full_name, message.from_user.username)

    # Проверка блокировки
    if is_user_blocked(user_id):
        _hm_lang = get_user_language(user_id) or "ru"
        await message.answer(
            '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> Your account has been blocked. Please contact the administrator.'
            if _hm_lang == "en"
            else '<tg-emoji emoji-id="5240241223632954241">❌</tg-emoji> Ваш аккаунт заблокирован. Обратитесь к администратору.'
        )
        return

    # Мой кабинет - открывает Mini App
    if message.text == "📱 Мой кабинет":
        mini_app_url_with_user = f"{MINI_APP_URL}?user_id={user_id}"
        await message.answer(
            "📱 Открываю ваш кабинет...",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [
                        KeyboardButton(
                            text=" Мой кабинет",
                            style="success",
                            icon_custom_emoji_id="5282843764451195532",
                            web_app=WebAppInfo(url=mini_app_url_with_user),
                        )
                    ],
                    [
                        KeyboardButton(
                            text=" Подписка",
                            style="primary",
                            icon_custom_emoji_id="5438496463044752972",
                        ),
                        KeyboardButton(
                            text=" Поддержка",
                            style="primary",
                            icon_custom_emoji_id="5443038326535759644",
                        ),
                    ],
                    [
                        KeyboardButton(
                            text=" Инструкцыя",
                            style="danger",
                            icon_custom_emoji_id="5222444124698853913",
                        ),
                        KeyboardButton(
                            text=" Язык/Language",
                            style="danger",
                            icon_custom_emoji_id="5447410659077661506",
                        ),
                    ],
                    [
                        KeyboardButton(
                            text="📤 Экспорт",
                            style="success",
                            icon_custom_emoji_id="5445355530111437729",
                        )
                    ],
                ],
                resize_keyboard=True,
            ),
        )

    # Подписка
    elif message.text == "💳 Подписка":
        await message.answer(
            "💳 <b>Информация о подписке</b>\n\n"
            "Вам доступен <b>7-дневный бесплатный пробный период</b>!\n\n"
            "После окончания пробного периода вы можете приобрести подписку.",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [
                        KeyboardButton(
                            text="📱 Мой кабинет",
                            web_app=WebAppInfo(url=f"{MINI_APP_URL}?user_id={user_id}"),
                        )
                    ],
                    [
                        KeyboardButton(text="💳 Подписка"),
                        KeyboardButton(text="💬 Поддержка"),
                    ],
                    [
                        KeyboardButton(text="📖 Инструкцыя"),
                        KeyboardButton(text="🌐 Язык/Language"),
                    ],
                    [KeyboardButton(text="📤 Экспорт")],
                ],
                resize_keyboard=True,
            ),
        )

    # Поддержка
    elif message.text == "💬 Поддержка":
        await message.answer(
            "💬 <b>Служба поддержки</b>\n\n"
            "По всем вопросам обращайтесь: @support\n\n"
            "Мы готовы помочь вам 24/7!",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [
                        KeyboardButton(
                            text="📱 Мой кабинет",
                            web_app=WebAppInfo(url=f"{MINI_APP_URL}?user_id={user_id}"),
                        )
                    ],
                    [
                        KeyboardButton(text="💳 Подписка"),
                        KeyboardButton(text="💬 Поддержка"),
                    ],
                    [
                        KeyboardButton(text="📖 Инструкцыя"),
                        KeyboardButton(text="🌐 Язык/Language"),
                    ],
                    [KeyboardButton(text="📤 Экспорт")],
                ],
                resize_keyboard=True,
            ),
        )

    # Инструкцыя
    elif message.text == "📖 Инструкцыя":
        await message.answer(
            "📖 <b>Инструкция по использованию бота</b>\n\n"
            '1. Нажмите "Мой кабинет" для управления финансами\n'
            "2. Добавляйте доходы и расходы\n"
            "3. Создавайте категории\n"
            "4. Устанавливайте напоминания\n"
            "5. Экспортируйте данные\n\n"
            "Приятного использования!",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [
                        KeyboardButton(
                            text=" Мой кабинет",
                            style="success",
                            icon_custom_emoji_id="5282843764451195532",
                            web_app=WebAppInfo(url=mini_app_url_with_user),
                        )
                    ],
                    [
                        KeyboardButton(
                            text=" Подписка",
                            style="primary",
                            icon_custom_emoji_id="5438496463044752972",
                        ),
                        KeyboardButton(
                            text=" Поддержка",
                            style="primary",
                            icon_custom_emoji_id="5443038326535759644",
                        ),
                    ],
                    [
                        KeyboardButton(
                            text=" Инструкцыя",
                            style="danger",
                            icon_custom_emoji_id="5222444124698853913",
                        ),
                        KeyboardButton(
                            text=" Язык/Language",
                            style="danger",
                            icon_custom_emoji_id="5447410659077661506",
                        ),
                    ],
                    [
                        KeyboardButton(
                            text="📤 Экспорт",
                            style="success",
                            icon_custom_emoji_id="5445355530111437729",
                        )
                    ],
                ],
                resize_keyboard=True,
            ),
        )

    # Язык/Language
    elif message.text == "🌐 Язык/Language":
        # Показываем клавиатуру выбора языка
        language_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=" Русский",
                        callback_data="ru",
                        icon_custom_emoji_id="5398017006165305287",
                    ),
                    InlineKeyboardButton(
                        text=" English",
                        callback_data="en",
                        icon_custom_emoji_id="5474446335744680905",
                    ),
                ]
            ]
        )
        await message.answer(
            '<tg-emoji emoji-id="5447410659077661506">🌐</tg-emoji> Выберите язык:\nChoose your language:',
            reply_markup=language_keyboard,
        )

    # Экспорт
    elif message.text == "📤 Экспорт":
        await message.answer(
            '<tg-emoji emoji-id="5395755469660762251">📤</tg-emoji> <b>Выберите период для выгрузки данных в <u>.XLSX</u></b>',
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [
                        KeyboardButton(
                            text=" Мой кабинет",
                            style="success",
                            icon_custom_emoji_id="5282843764451195532",
                            web_app=WebAppInfo(url=mini_app_url_with_user),
                        )
                    ],
                    [
                        KeyboardButton(
                            text=" Подписка",
                            style="primary",
                            icon_custom_emoji_id="5438496463044752972",
                        ),
                        KeyboardButton(
                            text=" Поддержка",
                            style="primary",
                            icon_custom_emoji_id="5443038326535759644",
                        ),
                    ],
                    [
                        KeyboardButton(
                            text=" Инструкцыя",
                            style="danger",
                            icon_custom_emoji_id="5222444124698853913",
                        ),
                        KeyboardButton(
                            text=" Язык/Language",
                            style="danger",
                            icon_custom_emoji_id="5447410659077661506",
                        ),
                    ],
                    [
                        KeyboardButton(
                            text="📤 Экспорт",
                            style="success",
                            icon_custom_emoji_id="5445355530111437729",
                        )
                    ],
                ],
                resize_keyboard=True,
            ),
        )


import psycopg2
from psycopg2.extras import RealDictCursor
import threading
from flask import Flask, render_template, request, jsonify
import json
from datetime import datetime, timezone, timedelta
import pytz
import requests as http_requests

app = Flask(__name__)

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://nytro_lone:0P1vnZDizoT3bfMqV7CnYT1AAiDE688o@dpg-d6lr2r24d50c73cg3js0-a.oregon-postgres.render.com/nytro_db",
)


def get_db_connection():
    conn = psycopg2.connect(DATABASE_URL)
    return conn


MONTH_NAMES_RU = [
    "",
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
]

MONTH_NAMES_EN = [
    "",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

MONTH_ABBR_RU = [
    "",
    "янв.",
    "февр.",
    "марта",
    "апр.",
    "мая",
    "июня",
    "июля",
    "авг.",
    "сент.",
    "окт.",
    "нояб.",
    "дек.",
]

MONTH_ABBR_EN = [
    "",
    "Jan.",
    "Feb.",
    "Mar.",
    "Apr.",
    "May",
    "Jun.",
    "Jul.",
    "Aug.",
    "Sep.",
    "Oct.",
    "Nov.",
    "Dec.",
]

MONTH_DAYS = [0, 31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


def is_leap_year(year):
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)


def days_in_month(year, month):
    if month == 2 and is_leap_year(year):
        return 29
    return MONTH_DAYS[month]


def parse_finance_date(date_str):
    """Парсит дату из формата DD.MM.YYYY или D.M.YYYY в объект date"""
    if not date_str:
        return None
    try:
        parts = date_str.strip().split(".")
        if len(parts) == 3:
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
            return datetime(year, month, day).date()
    except Exception:
        pass
    return None


def format_date_ru(year, month, day):
    """Форматирует дату в русском стиле: 5 янв. 2026 г."""
    return f"{day} {MONTH_ABBR_RU[month]} {year} г."


def format_date_en(year, month, day):
    """Formats date in English style: Jan. 5, 2026"""
    return f"{MONTH_ABBR_EN[month]} {day}, {year}"


def get_months_with_operations(user_id):
    """Возвращает список (year, month) месяцев, в которых есть хоть одна операция"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT date FROM finance WHERE user_id = %s", (user_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    months = set()
    for row in rows:
        d = parse_finance_date(row[0])
        if d:
            months.add((d.year, d.month))

    return sorted(months, reverse=True)


def get_finance_in_range(user_id, start_date, end_date):
    """Возвращает все операции пользователя в диапазоне дат (включительно)"""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        "SELECT id, type, amount, category, date, created_at FROM finance WHERE user_id = %s ORDER BY created_at ASC",
        (user_id,),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    result = []
    for row in rows:
        d = parse_finance_date(row["date"])
        if d and start_date <= d <= end_date:
            result.append(dict(row))
    return result


def generate_export_xlsx(records):
    """Генерирует XLSX файл из списка операций, возвращает bytes"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"

    headers = ["ID", "Тип", "Сумма", "Категория", "Дата"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4A90D9")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in records:
        type_label = "Доход" if r["type"] == "income" else "Расход"
        ws.append([r["id"], type_label, r["amount"], r["category"], r["date"]])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_export_months_keyboard(user_id, lang="ru"):
    """Строит инлайн-клавиатуру с месяцами и кнопкой произвольных дат"""
    months = get_months_with_operations(user_id)
    buttons = []
    month_names = MONTH_NAMES_EN if lang == "en" else MONTH_NAMES_RU
    for year, month in months:
        label = f"{month_names[month]} {year}"
        buttons.append(
            [InlineKeyboardButton(text=label, callback_data=f"expm_{year}_{month}")]
        )
    if lang == "en":
        buttons.append(
            [
                InlineKeyboardButton(
                    text=" Specify custom dates",
                    callback_data="expcustom",
                    icon_custom_emoji_id="6039779802741739617",
                )
            ]
        )
        buttons.append(
            [
                InlineKeyboardButton(
                    text=" Back",
                    callback_data="back_to_menu",
                    icon_custom_emoji_id="5352759161945867747",
                )
            ]
        )
    else:
        buttons.append(
            [
                InlineKeyboardButton(
                    text=" Указать произвольные даты",
                    callback_data="expcustom",
                    icon_custom_emoji_id="6039779802741739617",
                )
            ]
        )
        buttons.append(
            [
                InlineKeyboardButton(
                    text=" Назад",
                    callback_data="back_to_menu",
                    icon_custom_emoji_id="5352759161945867747",
                )
            ]
        )
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def build_export_confirm_keyboard(start_str, end_str, lang="ru"):
    """Строит клавиатуру подтверждения экспорта"""
    if lang == "en":
        return InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=" Export",
                        callback_data=f"expdo_{start_str}_{end_str}",
                        icon_custom_emoji_id="6021534008574217802",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="Cancel",
                        callback_data="expcancel",
                        icon_custom_emoji_id="5807626765874499116",
                    )
                ],
            ]
        )
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=" Экспортировать",
                    callback_data=f"expdo_{start_str}_{end_str}",
                    icon_custom_emoji_id="6021534008574217802",
                )
            ],
            [
                InlineKeyboardButton(
                    text="Отменить",
                    callback_data="expcancel",
                    icon_custom_emoji_id="5807626765874499116",
                )
            ],
        ]
    )


async def check_channel_subscription(user_id: int) -> bool:
    """Проверка подписки на все каналы"""
    try:
        for channel_username in REQUIRED_CHANNELS.values():
            chat_member = await bot.get_chat_member(
                chat_id=channel_username, user_id=user_id
            )
            if chat_member.status not in ["member", "administrator", "creator"]:
                return False
        return True
    except Exception as e:
        print(f"Error checking subscription: {e}")
        return False


def get_unsubscribed_channels(user_id: int) -> list:
    """Получение списка неподписанных каналов"""
    unsubscribed = []
    try:
        loop = asyncio.get_event_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    for channel_name, channel_username in REQUIRED_CHANNELS.items():
        try:
            chat_member = loop.run_until_complete(
                bot.get_chat_member(chat_id=channel_username, user_id=user_id)
            )
            if chat_member.status not in ["member", "administrator", "creator"]:
                unsubscribed.append((channel_name, channel_username))
        except Exception as e:
            print(f"Error checking channel {channel_username}: {e}")
            unsubscribed.append((channel_name, channel_username))
    return unsubscribed


def get_user_language(user_id: int) -> str:
    """Получение языка пользователя из БД"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT language FROM user_data WHERE user_id = %s", (user_id,))
    result = cur.fetchone()
    cur.close()
    conn.close()
    if result and result[0]:
        return result[0]
    return None  # Возвращаем None если язык не выбран


def is_user_blocked(user_id: int) -> bool:
    """Проверка, заблокирован ли пользователь"""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT is_blocked FROM user_data WHERE user_id = %s", (user_id,))
        result = cur.fetchone()
        return bool(result and result[0])
    except Exception:
        return False
    finally:
        cur.close()
        conn.close()


def save_user_info(user_id: int, full_name: str, username: str = None):
    """Сохраняет имя и username пользователя из Telegram в БД"""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO user_data (user_id, full_name, username)
            VALUES (%s, %s, %s)
            ON CONFLICT(user_id) DO UPDATE SET
                full_name = CASE WHEN EXCLUDED.full_name IS NOT NULL AND EXCLUDED.full_name != '' THEN EXCLUDED.full_name ELSE user_data.full_name END,
                username  = CASE WHEN EXCLUDED.username  IS NOT NULL AND EXCLUDED.username  != '' THEN EXCLUDED.username  ELSE user_data.username  END
            """,
            (user_id, full_name, username),
        )
        conn.commit()
    except Exception:
        pass
    finally:
        cur.close()
        conn.close()


def set_user_language(user_id: int, language: str):
    """Сохранение языка пользователя в БД"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO user_data (user_id, language)
        VALUES (%s, %s)
        ON CONFLICT(user_id) DO UPDATE SET language = EXCLUDED.language
    """,
        (user_id, language),
    )
    conn.commit()
    cur.close()
    conn.close()


def init_trial_subscription(user_id: int):
    """Выдаёт пробную 7-дневную подписку при первом входе"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT trial_given FROM user_data WHERE user_id = %s", (user_id,))
    row = cur.fetchone()
    if row and not row[0]:
        trial_end = datetime.now(timezone.utc) + timedelta(days=7)
        cur.execute(
            """
            UPDATE user_data
            SET subscription_end = %s, trial_given = TRUE, sub_notified = FALSE
            WHERE user_id = %s
        """,
            (trial_end, user_id),
        )
        conn.commit()
    cur.close()
    conn.close()


def extend_subscription(user_id: int, days: int):
    """Продлевает подписку пользователя на указанное количество дней"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT subscription_end FROM user_data WHERE user_id = %s", (user_id,))
    row = cur.fetchone()
    now = datetime.now(timezone.utc)
    if row and row[0]:
        current_end = row[0]
        if current_end.tzinfo is None:
            current_end = current_end.replace(tzinfo=timezone.utc)
        base = max(current_end, now)
    else:
        base = now
    new_end = base + timedelta(days=days)
    cur.execute(
        """
        UPDATE user_data
        SET subscription_end = %s, sub_notified = FALSE
        WHERE user_id = %s
    """,
        (new_end, user_id),
    )
    conn.commit()
    cur.close()
    conn.close()
    return new_end


def get_subscription_info(user_id: int):
    """Возвращает информацию о подписке пользователя"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT subscription_end FROM user_data WHERE user_id = %s", (user_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row or not row[0]:
        return {"active": False, "subscription_end": None}
    sub_end = row[0]
    if sub_end.tzinfo is None:
        sub_end = sub_end.replace(tzinfo=timezone.utc)
    now = datetime.now(timezone.utc)
    active = sub_end > now
    return {
        "active": active,
        "subscription_end": sub_end.strftime("%d.%m.%Y"),
    }


def init_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_data (
            user_id BIGINT PRIMARY KEY,
            username TEXT,
            full_name TEXT,
            language TEXT DEFAULT 'ru',
            stats TEXT,
            plans TEXT,
            settings TEXT,
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Добавляем колонку language, если она не существует (для существующих БД)
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS language TEXT DEFAULT 'ru'
    """)

    # Добавляем колонку created_at, если она не существует (для существующих БД)
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    """)

    # Добавляем колонки подписки
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS subscription_end TIMESTAMP
    """)
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS trial_given BOOLEAN DEFAULT FALSE
    """)
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS sub_notified BOOLEAN DEFAULT FALSE
    """)

    # Добавляем колонки блокировки и валюты
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS is_blocked BOOLEAN DEFAULT FALSE
    """)
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS currency TEXT DEFAULT 'rub'
    """)
    # Add notification_schedule_enabled column for 08:00/20:00 notifications
    cur.execute("""
        ALTER TABLE user_data ADD COLUMN IF NOT EXISTS notification_schedule_enabled BOOLEAN DEFAULT FALSE
    """)

    # Create notifications table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            name TEXT NOT NULL,
            datetime TEXT NOT NULL,
            timezone TEXT NOT NULL,
            type TEXT DEFAULT 'reminder',
            sent_notifications TEXT DEFAULT '[]',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES user_data(user_id)
        )
    """)

    # Create finance table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS finance (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            type TEXT NOT NULL,
            amount INTEGER NOT NULL,
            category TEXT NOT NULL,
            date TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES user_data(user_id)
        )
    """)

    # Create categories table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            name TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#6C5CE7',
            type TEXT NOT NULL DEFAULT 'expense',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES user_data(user_id)
        )
    """)
    # Add type column if it doesn't exist (for existing tables)
    try:
        cur.execute(
            "ALTER TABLE categories ADD COLUMN IF NOT EXISTS type TEXT NOT NULL DEFAULT 'expense'"
        )
    except Exception:
        pass

    # Create savings_goals table for kopilki
    cur.execute("""
        CREATE TABLE IF NOT EXISTS savings_goals (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL UNIQUE,
            goals_json TEXT DEFAULT '[]',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES user_data(user_id) ON DELETE CASCADE
        )
    """)

    conn.commit()
    cur.close()
    conn.close()


init_db()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/settings")
def settings():
    return render_template("settings.html")


@app.route("/money")
def money():
    return render_template("money.html")


@app.route("/admin")
def admin():
    return render_template("admin.html")


@app.route("/api/admin/stats")
def admin_stats():
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Total users
        cur.execute("SELECT COUNT(*) FROM user_data")
        total_users = cur.fetchone()[0]

        # Users for today
        cur.execute(
            "SELECT COUNT(*) FROM user_data WHERE DATE(created_at) = CURRENT_DATE"
        )
        day_users = cur.fetchone()[0]

        # Active subscriptions
        cur.execute("SELECT COUNT(*) FROM user_data WHERE subscription_end > NOW()")
        active_subs = cur.fetchone()[0]

        # Total savings (piggy banks) - count users with savings_goals
        cur.execute(
            "SELECT COUNT(*) FROM savings_goals WHERE goals_json != '[]' AND goals_json IS NOT NULL"
        )
        total_savings = cur.fetchone()[0]

        # Total plans - count plans in savings_goals
        cur.execute("SELECT goals_json FROM savings_goals")
        rows = cur.fetchall()
        total_plans = 0
        for row in rows:
            try:
                goals = json.loads(row[0]) if row[0] else []
                total_plans += len(goals)
            except:
                pass

        # Total notifications
        cur.execute("SELECT COUNT(*) FROM notifications")
        total_notifications = cur.fetchone()[0]

        # Total finance records
        cur.execute("SELECT COUNT(*) FROM finance")
        total_finance = cur.fetchone()[0]

        # Enabled notifications - count users with notification_schedule_enabled = true
        cur.execute(
            "SELECT COUNT(*) FROM user_data WHERE notification_schedule_enabled = TRUE"
        )
        enabled_notifications = cur.fetchone()[0]

        return jsonify(
            {
                "total_users": total_users,
                "day_users": day_users,
                "active_subs": active_subs,
                "total_savings": total_savings,
                "total_plans": total_plans,
                "total_notifications": total_notifications,
                "total_finance": total_finance,
                "enabled_notifications": enabled_notifications,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/admin/users")
def admin_users():
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT u.user_id, u.username, u.full_name, u.language, u.subscription_end,
                   u.created_at, u.is_blocked, u.currency,
                   (SELECT COUNT(*) FROM finance f WHERE f.user_id = u.user_id) AS finance_count
            FROM user_data u
            ORDER BY u.created_at DESC
        """)
        rows = cur.fetchall()
        users = []
        import json as _json

        now_utc = __import__("datetime").datetime.utcnow()
        for row in rows:
            sub_end = row[4]
            sub_active = bool(sub_end and sub_end > now_utc)
            # Count savings goals
            savings_count = 0
            try:
                cur2 = conn.cursor()
                cur2.execute(
                    "SELECT goals_json FROM savings_goals WHERE user_id = %s", (row[0],)
                )
                sg_row = cur2.fetchone()
                cur2.close()
                if sg_row and sg_row[0]:
                    parsed = _json.loads(sg_row[0])
                    savings_count = len(parsed) if isinstance(parsed, list) else 0
            except Exception:
                savings_count = 0
            users.append(
                {
                    "user_id": row[0],
                    "username": row[1],
                    "full_name": row[2],
                    "language": row[3],
                    "subscription_end": str(row[4]) if row[4] else None,
                    "sub_active": sub_active,
                    "created_at": str(row[5]) if row[5] else None,
                    "is_blocked": bool(row[6]),
                    "currency": row[7] or "rub",
                    "finance_count": row[8] or 0,
                    "savings_count": savings_count,
                }
            )
        return jsonify(users)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/admin/user/<int:user_id>/update", methods=["POST"])
def admin_update_user(user_id):
    data = request.json
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        full_name = data.get("full_name")
        currency = data.get("currency")
        cur.execute(
            """UPDATE user_data SET full_name = %s, currency = %s WHERE user_id = %s""",
            (full_name, currency, user_id),
        )
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)})
    finally:
        cur.close()
        conn.close()


@app.route("/api/admin/user/<int:user_id>/subscription", methods=["POST"])
def admin_manage_subscription(user_id):
    action = request.args.get("action", "grant")
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        if action == "remove":
            cur.execute(
                "UPDATE user_data SET subscription_end = NULL, sub_notified = FALSE WHERE user_id = %s",
                (user_id,),
            )
            conn.commit()
            return jsonify({"success": True})
        elif action == "grant":
            days = int(request.args.get("days", 30))
            if days <= 0:
                return jsonify(
                    {
                        "success": False,
                        "message": "Количество дней должно быть больше 0",
                    }
                )
            cur.close()
            conn.close()
            new_end = extend_subscription(user_id, days)
            return jsonify(
                {
                    "success": True,
                    "subscription_end": new_end.strftime("%Y-%m-%dT%H:%M:%S")
                    if new_end
                    else None,
                }
            )
        else:
            return jsonify({"success": False, "message": "Неизвестное действие"})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({"success": False, "message": str(e)})
    finally:
        try:
            cur.close()
            conn.close()
        except Exception:
            pass


@app.route("/api/admin/user/<int:user_id>/block", methods=["POST"])
def admin_block_user(user_id):
    action = request.args.get("action", "block")
    is_blocked = action == "block"
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE user_data SET is_blocked = %s WHERE user_id = %s",
            (is_blocked, user_id),
        )
        conn.commit()
        return jsonify({"success": True, "is_blocked": is_blocked})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)})
    finally:
        cur.close()
        conn.close()


@app.route("/api/admin/user/<int:user_id>/delete", methods=["POST"])
def admin_delete_user(user_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM finance WHERE user_id = %s", (user_id,))
        cur.execute("DELETE FROM notifications WHERE user_id = %s", (user_id,))
        cur.execute("DELETE FROM categories WHERE user_id = %s", (user_id,))
        cur.execute("DELETE FROM savings_goals WHERE user_id = %s", (user_id,))
        cur.execute("DELETE FROM user_data WHERE user_id = %s", (user_id,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)})
    finally:
        cur.close()
        conn.close()


@app.route("/api/save_savings/<int:user_id>", methods=["POST"])
def save_savings(user_id):
    data = request.json
    goals_json = data.get("goals_json", "[]")
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO savings_goals (user_id, goals_json, updated_at)
            VALUES (%s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (user_id) DO UPDATE SET goals_json = EXCLUDED.goals_json, updated_at = CURRENT_TIMESTAMP
        """,
            (user_id, goals_json),
        )
        conn.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/get_savings/<int:user_id>")
def get_savings(user_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT goals_json FROM savings_goals WHERE user_id = %s", (user_id,)
        )
        row = cur.fetchone()
        goals_json = row[0] if row else "[]"
        return jsonify({"goals_json": goals_json})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/save_currency/<int:user_id>", methods=["POST"])
def save_currency(user_id):
    data = request.json
    currency = data.get("currency", "rub")
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE user_data SET currency = %s WHERE user_id = %s", (currency, user_id)
        )
        conn.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/sync_profile", methods=["POST"])
def sync_profile():
    """Сохраняет только имя и username пользователя из Mini App при загрузке"""
    data = request.json
    user_id = data.get("user_id")
    if not user_id:
        return jsonify({"error": "No user_id"}), 400
    full_name = data.get("full_name", "").strip() or None
    username = data.get("username", "").strip() or None
    save_user_info(user_id, full_name, username)
    return jsonify({"status": "ok"})


@app.route("/api/set_language/<int:user_id>/<lang>", methods=["POST"])
def set_user_language_api(user_id, lang):
    """Устанавливает язык пользователя из Mini App и обновляет в боте"""
    if lang not in ("ru", "en"):
        return jsonify({"error": "Invalid language"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE user_data SET language = %s WHERE user_id = %s", (lang, user_id)
        )
        conn.commit()
        return jsonify({"status": "ok", "language": lang})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/toggle_notifications/<int:user_id>", methods=["POST"])
def toggle_notifications(user_id):
    """Включает/выключает уведомления в 08:00 и 20:00 и отправляет сообщение в бот"""
    data = request.json
    enabled = data.get("enabled", False)

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Update the database
        cur.execute(
            "UPDATE user_data SET notification_schedule_enabled = %s WHERE user_id = %s",
            (enabled, user_id),
        )
        conn.commit()

        # Get user's language
        cur.execute("SELECT language FROM user_data WHERE user_id = %s", (user_id,))
        row = cur.fetchone()
        user_lang = row[0] if row else "ru"

        # Build message
        if enabled:
            if user_lang == "en":
                message = "🔔 <b>Notifications enabled!</b>\n\nYou will receive daily reminders at <b>08:00</b> and <b>20:00</b> to check your finances."
            else:
                message = "🔔 <b>Уведомления включены!</b>\n\nВы будете получать ежедневные напоминания в <b>08:00</b> и <b>20:00</b> для проверки ваших финансов."
        else:
            if user_lang == "en":
                message = "🔕 <b>Notifications disabled!</b>\n\nYou will no longer receive daily reminders at 08:00 and 20:00."
            else:
                message = "🔕 <b>Уведомления выключены!</b>\n\nВы больше не будете получать ежедневные напоминания в 08:00 и 20:00."

        # Send message in a separate thread with its own event loop
        def send_message_thread():
            try:
                from aiogram import Bot
                import asyncio

                async def _send():
                    bot_thread = Bot(token=BOT_TOKEN)
                    try:
                        await bot_thread.send_message(
                            chat_id=user_id, text=message, parse_mode="HTML"
                        )
                    finally:
                        await bot_thread.session.close()

                asyncio.run(_send())
            except Exception as e:
                print(f"[toggle_notifications] Error sending message: {e}")

        import threading

        thread = threading.Thread(target=send_message_thread)
        thread.start()

        return jsonify({"status": "ok", "enabled": enabled})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/get_notification_status/<int:user_id>")
def get_notification_status(user_id):
    """Получает статус уведомлений пользователя"""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT notification_schedule_enabled FROM user_data WHERE user_id = %s",
            (user_id,),
        )
        row = cur.fetchone()
        enabled = row[0] if row else False
        return jsonify({"enabled": enabled})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/save_data", methods=["POST"])
def save_data():
    data = request.json
    user_id = data.get("user_id")
    if not user_id:
        return jsonify({"error": "No user_id provided"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO user_data (user_id, username, full_name, stats, plans, settings, last_updated)
        VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT(user_id) DO UPDATE SET
            username=EXCLUDED.username,
            full_name=EXCLUDED.full_name,
            stats=EXCLUDED.stats,
            plans=EXCLUDED.plans,
            settings=EXCLUDED.settings,
            last_updated=CURRENT_TIMESTAMP
    """,
        (
            user_id,
            data.get("username"),
            data.get("full_name"),
            data.get("stats"),
            data.get("plans"),
            data.get("settings"),
        ),
    )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"status": "success"})


@app.route("/api/get_data/<int:user_id>")
def get_data(user_id):
    # Проверка блокировки
    if is_user_blocked(user_id):
        return jsonify({"blocked": True, "error": "Ваш аккаунт заблокирован"}), 403

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT * FROM user_data WHERE user_id = %s", (user_id,))
    user = cur.fetchone()
    cur.close()
    conn.close()

    if user:
        return jsonify(
            {
                "user_id": user["user_id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "stats": user["stats"],
                "plans": user["plans"],
                "settings": user["settings"],
                "created_at": user["created_at"].strftime("%d.%m.%Y")
                if user["created_at"]
                else None,
            }
        )
    return jsonify({"error": "User not found"}), 404


@app.route("/api/subscription/<int:user_id>")
def get_subscription(user_id):
    init_trial_subscription(user_id)
    info = get_subscription_info(user_id)
    return jsonify(info)


@app.route("/api/update_settings", methods=["POST"])
def update_settings():
    data = request.json
    user_id = data.get("user_id")
    if not user_id:
        return jsonify({"error": "No user_id provided"}), 400

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT settings FROM user_data WHERE user_id = %s", (user_id,))
    row = cur.fetchone()
    existing = {}
    if row and row["settings"]:
        try:
            existing = json.loads(row["settings"])
        except Exception:
            existing = {}

    new_settings = data.get("settings", {})
    existing.update(new_settings)

    cur2 = conn.cursor()
    cur2.execute(
        "UPDATE user_data SET settings = %s WHERE user_id = %s",
        (json.dumps(existing), user_id),
    )
    conn.commit()
    cur.close()
    cur2.close()
    conn.close()
    return jsonify({"status": "success"})


@app.route("/api/save_notification", methods=["POST"])
def save_notification():
    data = request.json
    user_id = data.get("user_id")
    name = data.get("name")
    datetime_str = data.get("datetime")
    timezone_val = data.get("timezone", "UTC")
    notification_type = data.get("type", "reminder")

    if not user_id or not name or not datetime_str:
        return jsonify({"error": "Missing required fields"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO notifications (user_id, name, datetime, timezone, type, sent_notifications)
        VALUES (%s, %s, %s, %s, %s, '[]')
    """,
        (user_id, name, datetime_str, timezone_val, notification_type),
    )
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"status": "success", "message": "Notification saved"})


@app.route("/api/get_notifications/<int:user_id>")
def get_notifications(user_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        "SELECT * FROM notifications WHERE user_id = %s ORDER BY datetime ASC",
        (user_id,),
    )
    notifications = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify([dict(row) for row in notifications])


@app.route("/api/delete_notification/<int:notification_id>", methods=["DELETE"])
def delete_notification(notification_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM notifications WHERE id = %s", (notification_id,))
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"status": "success"})


# ==================== FINANCE API ====================


@app.route("/api/save_finance", methods=["POST"])
def save_finance():
    data = request.json
    user_id = data.get("user_id")
    finance_type = data.get("type")
    amount = data.get("amount")
    category = data.get("category")
    date = data.get("date")

    if not user_id or not finance_type or not amount or not category:
        return jsonify({"error": "Missing required fields"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO finance (user_id, type, amount, category, date)
        VALUES (%s, %s, %s, %s, %s)
    """,
        (user_id, finance_type, amount, category, date),
    )
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"status": "success"})


@app.route("/api/get_finance/<int:user_id>")
def get_finance(user_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        "SELECT * FROM finance WHERE user_id = %s ORDER BY created_at DESC", (user_id,)
    )
    finance = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify([dict(row) for row in finance])


@app.route("/api/update_finance", methods=["POST"])
def update_finance():
    data = request.json
    finance_id = data.get("id")
    user_id = data.get("user_id")
    finance_type = data.get("type")
    amount = data.get("amount")
    category = data.get("category")

    if not finance_id or not user_id or not finance_type or not amount or not category:
        return jsonify({"error": "Missing required fields"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE finance 
        SET type = %s, amount = %s, category = %s
        WHERE id = %s AND user_id = %s
    """,
        (finance_type, amount, category, finance_id, user_id),
    )
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"status": "success"})


@app.route("/api/delete_finance/<int:finance_id>", methods=["DELETE"])
def delete_finance(finance_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM finance WHERE id = %s", (finance_id,))
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"status": "success"})


@app.route("/api/get_categories/<int:user_id>")
def get_categories(user_id):
    cat_type = request.args.get("type")
    conn = get_db_connection()
    cur = conn.cursor()
    if cat_type:
        cur.execute(
            "SELECT id, name, color, type FROM categories WHERE user_id = %s AND type = %s ORDER BY id",
            (user_id, cat_type),
        )
    else:
        cur.execute(
            "SELECT id, name, color, type FROM categories WHERE user_id = %s ORDER BY id",
            (user_id,),
        )
    categories = cur.fetchall()
    cur.close()
    conn.close()

    result = []
    for cat in categories:
        result.append({"id": cat[0], "name": cat[1], "color": cat[2], "type": cat[3]})

    return jsonify(result)


@app.route("/api/add_category", methods=["POST"])
def add_category():
    data = request.json
    user_id = data.get("user_id")
    name = data.get("name")
    color = data.get("color", "#6C5CE7")
    category_type = data.get("type", "expense")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO categories (user_id, name, color, type)
        VALUES (%s, %s, %s, %s)
        RETURNING id
    """,
        (user_id, name, color, category_type),
    )
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True, "id": new_id})


@app.route("/api/update_category", methods=["POST"])
def update_category():
    data = request.json
    user_id = data.get("user_id")
    category_id = data.get("id")
    name = data.get("name")
    color = data.get("color")
    category_type = data.get("type")

    conn = get_db_connection()
    cur = conn.cursor()
    if category_type:
        cur.execute(
            """
            UPDATE categories 
            SET name = %s, color = %s, type = %s
            WHERE id = %s AND user_id = %s
        """,
            (name, color, category_type, category_id, user_id),
        )
    else:
        cur.execute(
            """
            UPDATE categories 
            SET name = %s, color = %s
            WHERE id = %s AND user_id = %s
        """,
            (name, color, category_id, user_id),
        )
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})


@app.route("/api/create_invoice", methods=["POST"])
def create_invoice():
    data = request.json
    payload = data.get("payload")

    plans = {
        "sub_1m": {
            "title": "Подписка на 1 месяц",
            "description": "30 дней доступа ко всем функциям NytroBC",
            "stars": 109,
        },
        "sub_3m": {
            "title": "Подписка на 3 месяца",
            "description": "90 дней доступа ко всем функциям NytroBC",
            "stars": 289,
        },
        "sub_6m": {
            "title": "Подписка на 6 месяцев",
            "description": "180 дней доступа ко всем функциям NytroBC",
            "stars": 539,
        },
        "sub_12m": {
            "title": "Подписка на 1 год",
            "description": "365 дней доступа ко всем функциям NytroBC",
            "stars": 849,
        },
    }

    if payload not in plans:
        return jsonify({"error": "Invalid plan"}), 400

    plan = plans[payload]

    response = http_requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/createInvoiceLink",
        json={
            "title": plan["title"],
            "description": plan["description"],
            "payload": payload,
            "provider_token": "",
            "currency": "XTR",
            "prices": [{"label": plan["title"], "amount": plan["stars"]}],
        },
    )

    result = response.json()
    if result.get("ok"):
        return jsonify({"url": result["result"]})
    else:
        return jsonify(
            {"error": result.get("description", "Не удалось создать счёт")}
        ), 500


@app.route("/api/delete_category", methods=["POST"])
def delete_category():
    data = request.json
    user_id = data.get("user_id")
    category_id = data.get("id")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM categories WHERE id = %s AND user_id = %s", (category_id, user_id)
    )
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})


# Reminder messages template
REMINDER_MESSAGES_RU = {
    60: '<tg-emoji emoji-id="5469947168523558652">⏰</tg-emoji> <b>Через 1 час</b> - {name}\n\nВремя на подготовку!',
    30: '<tg-emoji emoji-id="5382194935057372936">⏰</tg-emoji> <b>Через 30 минут</b> - {name}\n\nНе забудьте!',
    10: '<tg-emoji emoji-id="5458603043203327669">🔔</tg-emoji> <b>Через 10 минут</b> - {name}\n\nСкоро начнётся!',
    5: '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> <b>Через 5 минут</b> - {name}\n\nПриготовьтесь!',
    3: '<tg-emoji emoji-id="5395695537687123235">🚨</tg-emoji> <b>Срочно! Через 3 минуты</b> - {name}\n\nВставайте, начинаем!',
}

REMINDER_MESSAGES_EN = {
    60: '<tg-emoji emoji-id="5469947168523558652">⏰</tg-emoji> <b>In 1 hour</b> - {name}\n\nTime to get ready!',
    30: '<tg-emoji emoji-id="5382194935057372936">⏰</tg-emoji> <b>In 30 minutes</b> - {name}\n\nDon\'t forget!',
    10: '<tg-emoji emoji-id="5458603043203327669">🔔</tg-emoji> <b>In 10 minutes</b> - {name}\n\nStarting soon!',
    5: '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> <b>In 5 minutes</b> - {name}\n\nGet ready!',
    3: "<tg-emoji emoji-id=\"5395695537687123235\">🚨</tg-emoji> <b>Urgent! In 3 minutes</b> - {name}\n\nLet's go, we're starting!",
}

REMINDER_MESSAGES = REMINDER_MESSAGES_RU


async def send_reminder(
    user_id: int, notification_name: str, minutes_left: int, custom_message: str = None
):
    """Send reminder to user via Telegram"""
    if custom_message:
        message = custom_message
    else:
        _lang = get_user_language(user_id) or "ru"
        if _lang == "en":
            templates_ru = {
                60: '<tg-emoji emoji-id="5469947168523558652">⏰</tg-emoji> <b>In 1 hour</b> - {name} - Time to get ready!',
                30: '<tg-emoji emoji-id="5382194935057372936">⏰</tg-emoji> <b>In 30 minutes</b> - {name} - Don\'t forget!',
                10: '<tg-emoji emoji-id="5458603043203327669">🔔</tg-emoji> <b>In 10 minutes</b> - {name} - Starting soon!',
                5: '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> <b>In 5 minutes</b> - {name} - Get ready!',
                3: '<tg-emoji emoji-id="5395695537687123235">🚨</tg-emoji> <b>Urgent! In 3 minutes</b> - {name} - Let\'s go!',
            }
            message = templates_ru.get(
                minutes_left,
                '<tg-emoji emoji-id="5458603043203327669">🔔</tg-emoji> {name}',
            ).format(name=notification_name)
        else:
            templates = {
                60: '<tg-emoji emoji-id="5469947168523558652">⏰</tg-emoji> <b>Через 1 час</b> - {name} - Время на подготовку!',
                30: '<tg-emoji emoji-id="5382194935057372936">⏰</tg-emoji> <b>Через 30 минут</b> - {name} - Не забудьте!',
                10: '<tg-emoji emoji-id="5458603043203327669">🔔</tg-emoji> <b>Через 10 минут</b> - {name} - Скоро начнётся!',
                5: '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> <b>Через 5 минут</b> - {name} - Приготовьтесь',
                3: '<tg-emoji emoji-id="5395695537687123235">🚨</tg-emoji> <b>Срочно! Через 3 минуты</b> - {name} - Вставайте, начинаем!',
            }
            message = templates.get(
                minutes_left,
                '<tg-emoji emoji-id="5458603043203327669">🔔</tg-emoji> {name}',
            ).format(name=notification_name)

    try:
        await bot.send_message(user_id, message, parse_mode="HTML")
    except Exception as e:
        print(f"Error sending reminder: {e}")


def check_notifications():
    """Background task to check and send notifications"""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("SELECT * FROM notifications")
        notifications = cur.fetchall()
        now = datetime.now(timezone.utc)

        # Get or create a persistent event loop for the background thread
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

        for notification in notifications:
            try:
                notif_time = datetime.fromisoformat(
                    notification["datetime"].replace("Z", "+00:00")
                )
                diff = (notif_time - now).total_seconds() / 60
                minutes_until = int(diff)

                # Check which reminders should be sent
                sent = json.loads(notification["sent_notifications"] or "[]")

                for minutes in [60, 30, 10, 5, 3]:
                    # If we are within the window and haven't sent this specific reminder
                    if (
                        minutes_until <= minutes
                        and minutes_until > minutes - 2
                        and minutes not in sent
                    ):
                        # Use the persistent bot instance
                        asyncio.run_coroutine_threadsafe(
                            send_reminder(
                                notification["user_id"], notification["name"], minutes
                            ),
                            loop,
                        )

                        # Update sent notifications
                        sent.append(minutes)
                        cur.execute(
                            "UPDATE notifications SET sent_notifications = %s WHERE id = %s",
                            (json.dumps(sent), notification["id"]),
                        )
                        conn.commit()

                # Remove old notifications (past more than 5 minutes)
                if minutes_until < -5:
                    cur.execute(
                        "DELETE FROM notifications WHERE id = %s", (notification["id"],)
                    )
                    conn.commit()

            except Exception as e:
                print(f"Error processing notification {notification['id']}: {e}")
    finally:
        cur.close()
        conn.close()


def run_notification_checker():
    """Run notification checker every minute with its own event loop"""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    import time

    while True:
        try:
            check_notifications()
        except Exception as e:
            print(f"Error in notification checker: {e}")
        time.sleep(60)  # Check every minute


def run_flask():
    app.run(host="0.0.0.0", port=5000)


async def main():
    """Запуск бота"""
    dp.include_router(router)

    # Запускаем Flask в отдельном потоке
    threading.Thread(target=run_flask, daemon=True).start()

    # Get the main event loop
    main_loop = asyncio.get_running_loop()

    # Запускаем проверку уведомлений через asyncio.create_task вместо отдельного потока
    async def notification_scheduler():
        daily_sent = {"morning": None, "evening": None}
        while True:
            try:
                conn = get_db_connection()
                cur = conn.cursor(cursor_factory=RealDictCursor)
                try:
                    cur.execute("SELECT * FROM notifications")
                    notifications = cur.fetchall()
                    now = datetime.now(timezone.utc)

                    for notification in notifications:
                        try:
                            # Parse notification time and handle 'Z' suffix for UTC
                            dt_str = notification["datetime"]
                            if dt_str.endswith("Z"):
                                dt_str = dt_str[:-1] + "+00:00"
                            notif_time = datetime.fromisoformat(dt_str)

                            # Calculate minutes until notification in UTC
                            diff = (notif_time - now).total_seconds() / 60
                            minutes_until = int(diff)

                            # DEBUG LOG
                            print(
                                f"Checking notif {notification['id']} for user {notification['user_id']}: {minutes_until} min until {notif_time}"
                            )

                            sent = json.loads(
                                notification["sent_notifications"] or "[]"
                            )

                            for minutes in [60, 30, 10, 5, 3, 0]:
                                # If we are within the window
                                if (
                                    minutes_until <= minutes
                                    and minutes_until >= minutes - 1
                                    and minutes not in sent
                                ):
                                    # Send directly in the current loop
                                    print(
                                        f"!!! SENDING REMINDER {minutes} for {notification['name']} to {notification['user_id']}"
                                    )

                                    msg_text = ""
                                    if minutes == 0:
                                        msg_text = f"<tg-emoji emoji-id='5310278924616356636'>🎯</tg-emoji> <b>Время подошло!</b> - {notification['name']} - Начинаем прямо сейчас!"
                                    else:
                                        msg_text = None  # Use template in send_reminder

                                    await send_reminder(
                                        notification["user_id"],
                                        notification["name"],
                                        minutes,
                                        custom_message=msg_text,
                                    )

                                    sent.append(minutes)
                                    cur.execute(
                                        "UPDATE notifications SET sent_notifications = %s WHERE id = %s",
                                        (json.dumps(sent), notification["id"]),
                                    )
                                    conn.commit()

                            if minutes_until < -5:
                                cur.execute(
                                    "DELETE FROM notifications WHERE id = %s",
                                    (notification["id"],),
                                )
                                conn.commit()
                        except Exception as e:
                            print(
                                f"Error processing notification {notification['id']}: {e}"
                            )
                    # Ежедневные уведомления в 09:00 и 20:00 МСК (UTC+3: 06:00 и 17:00 UTC)
                    try:
                        msk_now = now + timedelta(hours=3)
                        today = msk_now.date()
                        current_hour = msk_now.hour
                        current_minute = msk_now.minute

                        DAILY_MSGS = {
                            "morning": {
                                "ru": '<b><u>Доброе утро!</u> <tg-emoji emoji-id="5469947168523558652">☀️</tg-emoji>\n<i>Не забудьте сегодня записать свои доходы и расходы, а также выполнить запланированные дела.</i>\nПусть день будет продуктивным и успешным! <tg-emoji emoji-id="5359785904535774578">💼</tg-emoji></b>',
                                "en": '<b><u>Good morning!</u> <tg-emoji emoji-id="5469947168523558652">☀️</tg-emoji>\n<i>Don\'t forget to record your income and expenses today, and complete your planned tasks.</i>\nMay the day be productive and successful! <tg-emoji emoji-id="5359785904535774578">💼</tg-emoji></b>',
                            },
                            "evening": {
                                "ru": '<b><u>Добрый вечер!</u> <tg-emoji emoji-id="5769143090103193926">🌙</tg-emoji>\n<i>Похоже, сегодня был продуктивный и успешный день — это отлично!</i> Не забудьте зафиксировать свои расходы и пополнения, чтобы держать финансы под контролем. <tg-emoji emoji-id="5224257782013769471">💰</tg-emoji></b>',
                                "en": '<b><u>Good evening!</u> <tg-emoji emoji-id="5769143090103193926">🌙</tg-emoji>\n<i>Looks like today was a productive and successful day — that\'s great!</i> Don\'t forget to record your expenses and income to keep your finances under control. <tg-emoji emoji-id="5224257782013769471">💰</tg-emoji></b>',
                            },
                        }
                        for period, hour in [("morning", 8), ("evening", 20)]:
                            if (
                                current_hour == hour
                                and current_minute == 0
                                and daily_sent[period] != today
                            ):
                                daily_sent[period] = today
                                cur_d = conn.cursor(cursor_factory=RealDictCursor)
                                cur_d.execute(
                                    "SELECT user_id, settings, language, notification_schedule_enabled FROM user_data"
                                )
                                all_users = cur_d.fetchall()
                                cur_d.close()
                                for u in all_users:
                                    try:
                                        # Check both the new column and old settings field
                                        should_send = u.get(
                                            "notification_schedule_enabled", False
                                        )
                                        if not should_send:
                                            s = json.loads(u["settings"] or "{}")
                                            should_send = s.get(
                                                "notifications_enabled", False
                                            )

                                        if should_send:
                                            u_lang = u.get("language") or "ru"
                                            msg_text = DAILY_MSGS[period].get(
                                                u_lang, DAILY_MSGS[period]["ru"]
                                            )
                                            await bot.send_message(
                                                chat_id=u["user_id"],
                                                text=msg_text,
                                                parse_mode="HTML",
                                            )
                                    except Exception as e_u:
                                        print(
                                            f"Error sending daily {period} notif to {u['user_id']}: {e_u}"
                                        )
                    except Exception as e_daily:
                        print(f"Error in daily notification check: {e_daily}")

                    # Проверяем истечение подписки
                    try:
                        cur2 = conn.cursor(cursor_factory=RealDictCursor)
                        cur2.execute(
                            """
                            SELECT user_id, language
                            FROM user_data
                            WHERE subscription_end IS NOT NULL
                              AND subscription_end <= %s
                              AND subscription_end >= %s
                              AND (sub_notified IS NULL OR sub_notified = FALSE)
                        """,
                            (now, now - timedelta(minutes=2)),
                        )
                        expired_users = cur2.fetchall()
                        for u in expired_users:
                            try:
                                lang = u.get("language") or "ru"
                                open_btn = InlineKeyboardMarkup(
                                    inline_keyboard=[
                                        [
                                            InlineKeyboardButton(
                                                text="Open"
                                                if lang == "en"
                                                else "Открыть",
                                                web_app=WebAppInfo(
                                                    url=f"{MINI_APP_URL}?user_id={u['user_id']}&page=settings"
                                                ),
                                            )
                                        ]
                                    ]
                                )
                                if lang == "en":
                                    msg = (
                                        '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> <b>Your subscription has expired!</b>\n\n'
                                        "Please renew it in the Settings section of our Mini App."
                                    )
                                else:
                                    msg = (
                                        '<tg-emoji emoji-id="5447644880824181073">⚠️</tg-emoji> <b>Ваша подписка закончилась!</b>\n\n'
                                        "Пожалуйста, продлите её в разделе Настройки нашего Mini App."
                                    )
                                await bot.send_message(
                                    chat_id=u["user_id"],
                                    text=msg,
                                    parse_mode="HTML",
                                    reply_markup=open_btn,
                                )
                                cur2.execute(
                                    "UPDATE user_data SET sub_notified = TRUE WHERE user_id = %s",
                                    (u["user_id"],),
                                )
                                conn.commit()
                            except Exception as e_inner:
                                print(
                                    f"Error sending sub expiry notif to {u['user_id']}: {e_inner}"
                                )
                        cur2.close()
                    except Exception as e_sub:
                        print(f"Error in subscription expiry check: {e_sub}")

                finally:
                    cur.close()
                    conn.close()
            except Exception as e:
                print(f"Error in notification scheduler: {e}")
            await asyncio.sleep(60)

    asyncio.create_task(notification_scheduler())

    # Удаляем webhook если есть
    await bot.delete_webhook(drop_pending_updates=True)

    # Запускаем polling
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
